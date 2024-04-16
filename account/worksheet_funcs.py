import os
from datetime import datetime
import calendar
import locale
from datetime import datetime, timedelta
from itertools import groupby

from django.conf import settings

import pymorphy3
from petrovich.main import Petrovich
from petrovich.enums import Case, Gender
from pytils import numeral

from openpyxl import load_workbook
from docxtpl import DocxTemplate


APP_PATH = os.path.dirname(os.path.abspath(__file__))


def parse_date(date_str):
    return datetime.strptime(date_str, '%d.%m.%Y')

def delta_days(date1, date2):
    return (date2 - date1).days

def group_consecutive_dates(data):
    sorted_data = sorted(data, key=lambda x: parse_date(x['date']))  # Сортировка по дате
    groups = []
    temp_group = [sorted_data[0]]

    for prev, current in zip(sorted_data, sorted_data[1:]):
        if delta_days(parse_date(prev['date']), parse_date(current['date'])) == 1:
            temp_group.append(current)
        else:
            groups.append(temp_group)
            temp_group = [current]
    groups.append(temp_group)  # Добавляем последнюю группу

    return groups


def get_employee_info(
    file_path: str,
    sheet_name: str,
    month: int,
    month_part: int,
    year: int,
    date_row: int,
    employee_col: int
) -> list:
    
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    final_data = []
    
    # Проходится по строке DATE_ROW и ищет числа, после чего берет координату последнего числа
    # для определения диапазона поиска в дальнейшем
    days = [
        {
            'value': ws.cell(row=date_row, column=c+1).value,
            'row': ws.cell(row=date_row, column=c+1).column,
        }
        for c in range(ws.max_column) if ws.cell(row=date_row, column=c+1).value != None
    ]
    
    # Последний день отчетного месяца
    _, month_last_day = calendar.monthrange(int(year), int(month))
    
    if int(month_part) == 1:
        first_row = next((i for i, d in enumerate(days) if d['value'] == 1), None)
        last_row = next((i for i, d in enumerate(days) if d['value'] == 15), None)
    elif int(month_part) == 2:
        first_row = next((i for i, d in enumerate(days) if d['value'] == 16), None)
        last_row = next((i for i, d in enumerate(days) if d['value'] == month_last_day), None)
        
    first_day = days[first_row]['row']
    last_day = days[last_row]['row']
    
    for r in range(len(ws['B'])):
        employee_cell = ws.cell(row=r+1, column=employee_col)
        post_cell = ws.cell(row=r+1, column=employee_col+1)
        employee_info = {
            'fullname': employee_cell.value,
            'post': post_cell.value,
            'overwork': None,
        }
        
        work_dates = []
        # Проверка окрашена ячейка или нет
        if employee_cell.fill.fgColor.value != '00000000':
            # Проходим диапазон до последнего рабочего дня last_day
            for i in range(first_day - 1, last_day):
                work_status_cell = ws.cell(row=employee_cell.row, column=i+1)
                work_hours_cell = ws.cell(row=employee_cell.row+1, column=i+1)
                work_date_cell = ws.cell(row=date_row, column=i+1)

                if work_status_cell.value == 'РВ':
                    work_dates.append({
                        'date': f'{work_date_cell.value}.{month:02}.{year}',
                        'hours': work_hours_cell.value
                    })
        
        if work_dates != []:
            grouped_dates = group_consecutive_dates(work_dates)
            employee_info['overwork'] = grouped_dates
        
        if employee_info['overwork'] is not None:
            final_data.append(employee_info)

    return final_data


def create_office_files(
    memo_path: str,
    order_path: str,
    order_plural_path: str,
    employee_info: list,
    chief: str
) -> None:    

    petrovich = Petrovich()
    
    worksheet_dir = os.path.join(APP_PATH, 'worksheet_docs')
    try:
        os.mkdir(os.path.join(worksheet_dir, 'raw_files'))
    except FileExistsError:
        pass
    
    # all_employees = []
    for employee in employee_info:        
        fullname = employee['fullname'].split(' ')
        
        normal_firstname = fullname[1]
        normal_middlename = fullname[2]
        normal_lastname = fullname[0]
        cased_firstname = petrovich.firstname(normal_firstname, Case.GENITIVE, Gender.MALE)
        cased_middlename = petrovich.firstname(normal_middlename, Case.GENITIVE, Gender.MALE)
        cased_lastname = petrovich.firstname(normal_lastname, Case.GENITIVE, Gender.MALE)
        dative_firstname = petrovich.firstname(normal_firstname, Case.DATIVE, Gender.MALE)
        dative_middlename = petrovich.firstname(normal_middlename, Case.DATIVE, Gender.MALE)
        dative_lastname = petrovich.firstname(normal_lastname, Case.DATIVE, Gender.MALE)
        
        cased_chief = petrovich.lastname(chief, Case.DATIVE, Gender.MALE)
        
        post = employee['post'].lower()
        if 'начальник участка' in post.lower():
            cased_post = petrovich.lastname(post.split()[0], Case.GENITIVE, Gender.MALE)
            cased_post += ' ' + ' '.join(post.split()[1:])
        else:
            cased_post = petrovich.lastname(post, Case.GENITIVE, Gender.MALE)

        context = {
            'normal_chief': chief,
            'cased_chief': cased_chief,
            'normal_fullname': f'{normal_lastname} {normal_firstname} {normal_middlename}',
            'cased_fullname': f'{cased_lastname} {cased_firstname} {cased_middlename}',
            'dative_fullname': f'{dative_lastname} {dative_firstname} {dative_middlename}',
            'normal_post': post,
            'cased_post': cased_post,
        }
        
        # формирование строки для приказа, если несколько дней
        for overwork_list in employee['overwork']:
            order_dates = ''
            date_obj = parse_date(overwork_list[0]['date'])
            
            # получения дня недели от числа
            memo_week = date_obj.weekday() + 1
            
            # если начало переработки выпадает на воскресенье то приказ делается от пятницы
            if memo_week == 7:
                memo_day = date_obj.day - 2
            else:
                memo_day = date_obj.day - 1
            
            for overwork in overwork_list:
                date = overwork['date']

                context['date'] = date
                
                locale.setlocale(locale.LC_TIME, 'ru_RU.UTF-8')
                morph = pymorphy3.MorphAnalyzer()                

                context['memo_day'] = memo_day
                
                # получение названия месяца
                month = calendar.month_name[date_obj.month]
                # склонение месяца
                cased_month = morph.parse(month)[0].word
                year = date_obj.year
                
                context['month'] = month
                context['year'] = year
                context['cased_month'] = cased_month
                
                # проверка на корректность заполнения часов
                if overwork['hours'] is not None:
                    hour = numeral.get_plural(overwork['hours'], 'час, часа, часов')
                else:
                    hour = numeral.get_plural(0, 'час, часа, часов')
                    
                context['hour'] = hour

                order_dates += f'- {date} на {hour}\n'
                
                # формирование служебной записки
                memo_doc = DocxTemplate(memo_path)
                memo_doc.render(context)
                memo_file = os.path.join(APP_PATH, f'worksheet_docs/raw_files/{normal_lastname} - {date}.docx')
                memo_doc.save(memo_file)
        
            context['order_dates'] = order_dates
            
            # формирование приказа
            if len(overwork_list) <= 1:
                order_doc = DocxTemplate(order_path)
            else:
                order_doc = DocxTemplate(order_plural_path)  
            order_doc.render(context)
            order_file = os.path.join(APP_PATH, f'worksheet_docs/raw_files/Приказ {normal_lastname} - {date}.docx')
            order_doc.save(order_file)


def main(form_data: dict) -> None:
    excel_path = os.path.join(settings.MEDIA_ROOT, form_data['file_name'])
    doc_path = os.path.join(APP_PATH, 'worksheet_docs/templates')
    
    employee_info = get_employee_info(
        file_path=excel_path,
        sheet_name=form_data['sheet_name'],
        month=form_data['month'],
        month_part=form_data['month_part'],
        year=form_data['year'],
        date_row=form_data['date_row'],
        employee_col=form_data['employee_col'],
    )
    create_office_files(
        memo_path=os.path.join(doc_path, 'office_memo.docx'),
        order_path=os.path.join(doc_path, 'office_order.docx'),
        order_plural_path=os.path.join(doc_path, 'office_order_plural.docx'),
        employee_info=employee_info,
        chief=form_data['chief'],
    )

